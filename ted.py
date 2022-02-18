import requests
from bs4 import BeautifulSoup
import os
import re
import lxml
from selenium import webdriver
from requests_html import HTMLSession
import xlsxwriter
from openpyxl import load_workbook

from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

head = {
    'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.0 x64; en-US; rv:1.9pre) Gecko/2008072421 Minefield/3.0.2pre'
}

PATH_TO_CHROME_SERVER = '/Users/linyaya/Desktop/a star/ted_scraper/chromedriver'

LANGS = ['en', 'zh-cn', 'ms', 'id']

SAVE_FILE = 'ted.xlsx'


def get_TED_talk_list(url, i):
    url_ted = url + str(i)
    response = requests.get(url_ted, params=head)
    html = response.text
    bs = BeautifulSoup(html, 'html.parser')
    talks_list = bs.find_all('div', attrs={'class': 'media__message'})

    talk_url_dic = {}
    for j in range(len(talks_list)):
        ted_a = talks_list[j].find_all(
            'a', attrs={'class': 'ga-link', 'data-ga-context': 'talks'})
        ted_url = 'https://www.ted.com' + ted_a[0]['href']
        talk_title = ted_a[0].text.strip()
        talk_url_dic[talk_title] = ted_url
    return talk_url_dic


def get_language_transcript(driver):
    soup = BeautifulSoup(driver.page_source, 'lxml')
    content_selector = '#maincontent aside div.w-full span.cursor-pointer.inline.hover\:bg-red-300.css-82uonn'
    content_ele_list = soup.select(content_selector)
    # print(content_ele_list)
    content = ''
    for content_ele in content_ele_list:
        content += content_ele.text.replace('\n', '')
    return content


def talk_write_into_file(title, url, driver, ws):
    url = url + '/transcript'
    driver.get(url)

    select_selector = 'select.rounded-sm.css-wbqv41'
    try:
        select = Select(driver.find_element_by_css_selector(select_selector))
    except:
        print('cannot find select element / cannot get transcript')
        return
    language_code_list = [option.get_attribute(
        'value') for option in select.options]

    if 'en' not in language_code_list:
        return
    has_other = False
    for lang_code in LANGS[1:]:
        if lang_code in language_code_list:
            has_other = True
            continue
    if not has_other:
        print('not have other language')
        return

    content_list = [title]
    for i, lang_code in enumerate(LANGS):
        content = ''
        if lang_code in language_code_list:
            select.select_by_value(lang_code)
            time.sleep(3)
            content = get_language_transcript(driver)
        content_list.append(content)

    ws.append(content_list)


if __name__ == '__main__':

    index = 0
    '''
    workbook = xlsxwriter.Workbook(SAVE_FILE)
    # Add style
    bold = workbook.add_format({'bold': True})
    regular = workbook.add_format({'bold': False})

    # Add sheet
    worksheet = workbook.add_worksheet('data' + str(index))
    # Write sheet head
    row0 = ["title",   "English", "Chinese", "Malay", "Indo"]
    for i in range(0, len(row0)):
        worksheet.write(0, i, row0[i], bold)
    workbook.close()
    '''

    wb = load_workbook(SAVE_FILE)
    # Select First Worksheet
    ws = wb.worksheets[index]

    # driver = webdriver.PhantomJS(executable_path='/Users/linyaya/Desktop/phantomjs-2.1.1-macosx/bin/phantomjs')
    driver = webdriver.Chrome(PATH_TO_CHROME_SERVER)

    # url = 'https://www.ted.com/talks/iseult_gillespie_the_myth_of_narcissus_and_echo'
    # talk_write_into_file('', url, driver, ws)

    base_url = 'https://www.ted.com/talks?page='
    start = int(input("input start page: "))
    end = int(input("input end page: "))

    for i in range(start, end+1):
        print('==> page: ' + str(i))
        talk_url_dic = get_TED_talk_list(base_url, i)
        for j, (title, url) in enumerate(talk_url_dic.items()):
            talk_write_into_file(title, url, driver, ws)
            print('====> page: ' + str(i) + ' no.' + str(j+1) +
                  ' write into file, video title: ' + title)
            wb.save(SAVE_FILE)
    driver.quit()
